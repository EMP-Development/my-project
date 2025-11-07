import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel転記ツール v3.6（除外行分離版・ログ修正）")
        self.root.geometry("1200x750")

        # 変数定義
        self.src_file = tk.StringVar()
        self.dst_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.src_key_col = tk.StringVar()
        self.dst_key_col = tk.StringVar()
        self.src_start_row = tk.IntVar(value=1)
        self.dst_start_row = tk.IntVar(value=1)
        self.src_exclude_rows = tk.StringVar(value="")
        self.dst_exclude_rows = tk.StringVar(value="")
        self.exec_mode = tk.StringVar(value="上書き")

        self.mappings = []

        self.build_ui()

    def build_ui(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)

        # === 元データ ===
        src_frame = ttk.LabelFrame(top_frame, text="元データ設定")
        src_frame.pack(side="left", fill="both", expand=True, padx=5)

        ttk.Button(src_frame, text="ファイル選択", command=lambda: self.select_file(self.src_file, self.src_sheet)).pack(pady=2)
        ttk.Entry(src_frame, textvariable=self.src_file, width=50).pack(pady=2)
        ttk.Label(src_frame, text="シート選択:").pack()
        self.src_sheet_cb = ttk.Combobox(src_frame, textvariable=self.src_sheet, state="readonly")
        self.src_sheet_cb.pack(pady=2)
        ttk.Label(src_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_key_col, width=10).pack(pady=2)
        ttk.Label(src_frame, text="開始行:").pack()
        ttk.Entry(src_frame, textvariable=self.src_start_row, width=5).pack(pady=2)
        ttk.Label(src_frame, text="除外行（例: 1,2,3）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_exclude_rows, width=15).pack(pady=2)

        # === 転記先データ ===
        dst_frame = ttk.LabelFrame(top_frame, text="転記先設定")
        dst_frame.pack(side="right", fill="both", expand=True, padx=5)

        ttk.Button(dst_frame, text="ファイル選択", command=lambda: self.select_file(self.dst_file, self.dst_sheet)).pack(pady=2)
        ttk.Entry(dst_frame, textvariable=self.dst_file, width=50).pack(pady=2)
        ttk.Label(dst_frame, text="シート選択:").pack()
        self.dst_sheet_cb = ttk.Combobox(dst_frame, textvariable=self.dst_sheet, state="readonly")
        self.dst_sheet_cb.pack(pady=2)
        ttk.Label(dst_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_key_col, width=10).pack(pady=2)
        ttk.Label(dst_frame, text="開始行:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_start_row, width=5).pack(pady=2)
        ttk.Label(dst_frame, text="除外行（例: 1,2,3）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_exclude_rows, width=15).pack(pady=2)

        # === 転記設定 ===
        mapping_frame = ttk.LabelFrame(self.root, text="転記設定（元列 → 転記先列）")
        mapping_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.mapping_container = ttk.Frame(mapping_frame)
        self.mapping_container.pack(fill="x", padx=5, pady=5)
        ttk.Button(mapping_frame, text="＋ 条件を追加", command=self.add_mapping).pack(anchor="e", padx=10, pady=5)
        self.add_mapping()

        # === 下部 ===
        bottom_frame = ttk.LabelFrame(self.root, text="実行設定")
        bottom_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(bottom_frame, text="実行モード:").pack(side="left")
        ttk.Radiobutton(bottom_frame, text="上書き", value="上書き", variable=self.exec_mode).pack(side="left", padx=5)
        ttk.Radiobutton(bottom_frame, text="既存保持", value="既存保持", variable=self.exec_mode).pack(side="left", padx=5)
        ttk.Radiobutton(bottom_frame, text="確認のみ", value="確認のみ", variable=self.exec_mode).pack(side="left", padx=5)

        ttk.Button(bottom_frame, text="転記実行", command=self.execute_mapping).pack(side="right", padx=10)

    # === ファイル選択 ===
    def select_file(self, var, sheet_var):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not file_path:
            return
        var.set(file_path)
        try:
            xl = pd.ExcelFile(file_path)
            if var == self.src_file:
                self.src_sheet_cb["values"] = xl.sheet_names
                if xl.sheet_names:
                    sheet_var.set(xl.sheet_names[0])
            else:
                self.dst_sheet_cb["values"] = xl.sheet_names
                if xl.sheet_names:
                    sheet_var.set(xl.sheet_names[0])
        except Exception as e:
            messagebox.showerror("エラー", f"シート取得失敗:\n{e}")

    def add_mapping(self):
        frame = ttk.Frame(self.mapping_container)
        frame.pack(fill="x", pady=2)

        src_col = tk.StringVar()
        dst_col = tk.StringVar()
        self.mappings.append((src_col, dst_col))

        ttk.Label(frame, text="元列:").pack(side="left")
        ttk.Entry(frame, textvariable=src_col, width=10).pack(side="left", padx=5)
        ttk.Label(frame, text="→ 転記先列:").pack(side="left")
        ttk.Entry(frame, textvariable=dst_col, width=10).pack(side="left", padx=5)

        if len(self.mappings) > 1:
            ttk.Button(frame, text="削除", command=lambda f=frame: self.remove_mapping(f)).pack(side="left", padx=5)

    def remove_mapping(self, frame):
        frame.destroy()

    def col_to_index(self, col):
        idx = 0
        for c in col:
            if 'A' <= c.upper() <= 'Z':
                idx = idx * 26 + (ord(c.upper()) - 64)
        return idx - 1

    # === 実行 ===
    def execute_mapping(self):
        try:
            src_path, dst_path = self.src_file.get(), self.dst_file.get()
            src_sheet, dst_sheet = self.src_sheet.get(), self.dst_sheet.get()
            src_key_col, dst_key_col = self.src_key_col.get().strip(), self.dst_key_col.get().strip()
            src_start, dst_start = self.src_start_row.get(), self.dst_start_row.get()
            mode = self.exec_mode.get()

            src_exclude = [int(x.strip()) for x in self.src_exclude_rows.get().split(",") if x.strip()]
            dst_exclude = [int(x.strip()) for x in self.dst_exclude_rows.get().split(",") if x.strip()]

            if not all([src_path, dst_path, src_sheet, dst_sheet, src_key_col, dst_key_col]):
                messagebox.showwarning("警告", "ファイル・シート・キー列が指定されていません。")
                return

            mappings = [(s.get().strip().upper(), d.get().strip().upper()) for s, d in self.mappings if s.get().strip() and d.get().strip()]
            if not mappings:
                messagebox.showwarning("警告", "転記設定がありません。")
                return

            # 読み込み（生データ）
            df_src = pd.read_excel(src_path, sheet_name=src_sheet, header=None)
            df_dst = pd.read_excel(dst_path, sheet_name=dst_sheet, header=None)
            wb_dst = load_workbook(dst_path)
            ws_dst = wb_dst[dst_sheet]

            # ログ準備（先頭にファイル・シート情報を明示）
            log_dir = os.path.join(os.path.expanduser("~/Desktop"), "コピペログ")
            os.makedirs(log_dir, exist_ok=True)
            log_name = datetime.now().strftime("%Y.%m.%d.%H.%M.%S.xlsx")
            log_path = os.path.join(log_dir, log_name)
            wb_log = Workbook()
            ws_log = wb_log.active
            ws_log.title = "転記ログ"

            # --- メタ情報（ファイル名・シート名）を明示する位置に出力 ---
            ws_log["A1"] = "転記元ファイル"
            ws_log["B2"] = os.path.basename(src_path)
            ws_log["B3"] = src_sheet
            ws_log["A4"] = "転記先ファイル"
            ws_log["B5"] = os.path.basename(dst_path)
            ws_log["B6"] = dst_sheet
            ws_log["C7"] = f"モード: {mode}"

            # ヘッダーは8行目から
            header = ["転記先行", "転記元行", "照合キー値"]
            for s, d in mappings:
                header += [f"転記先{d}(列名)", f"転記元{s}(列名)"]
            for i, h in enumerate(header):
                ws_log.cell(row=8, column=i + 1, value=h)

            # キー列インデックス
            src_key_idx = self.col_to_index(src_key_col)
            dst_key_idx = self.col_to_index(dst_key_col)

            # 転記先辞書（dst_key_val -> index）
            dst_dict = {}
            for i in range(dst_start - 1, len(df_dst)):
                if (i + 1) in dst_exclude:
                    continue
                try:
                    val = df_dst.iat[i, dst_key_idx]
                except Exception:
                    val = None
                if pd.notna(val):
                    dst_dict[val] = i

            # 実処理（srcを走査）
            log_row = 9
            for i in range(src_start - 1, len(df_src)):
                if (i + 1) in src_exclude:
                    continue
                try:
                    key_val = df_src.iat[i, src_key_idx]
                except Exception:
                    key_val = None
                if pd.isna(key_val):
                    continue
                if key_val not in dst_dict:
                    continue

                dst_i = dst_dict[key_val]
                if (dst_i + 1) in dst_exclude:
                    continue

                # Excel上の行番号を反映（1-based）
                dst_excel_row = dst_i + 1
                src_excel_row = i + 1

                log_values = [dst_excel_row, src_excel_row, key_val]

                for s, d in mappings:
                    s_idx = self.col_to_index(s)
                    d_idx = self.col_to_index(d)
                    src_val = df_src.iat[i, s_idx] if s_idx < len(df_src.columns) else None
                    dst_cell = ws_dst.cell(row=dst_excel_row, column=d_idx + 1)
                    dst_before = dst_cell.value
                    dst_after = dst_before

                    if mode == "上書き":
                        dst_after = src_val
                        dst_cell.value = src_val
                    elif mode == "既存保持":
                        if dst_before is None:
                            dst_after = src_val
                            dst_cell.value = src_val
                    elif mode == "確認のみ":
                        dst_after = src_val if (dst_before is None or dst_before != src_val) else dst_before

                    log_values += [dst_after, src_val]

                # ログ書き込み（8行目の下から）
                for col_idx, v in enumerate(log_values, start=1):
                    ws_log.cell(row=log_row, column=col_idx, value=v)
                log_row += 1

            # ログ保存と（モードによる）ワークブック保存
            wb_log.save(log_path)
            if mode != "確認のみ":
                wb_dst.save(dst_path)

            messagebox.showinfo("完了", f"転記完了（モード：{mode}）\nログ出力: {log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラー:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
