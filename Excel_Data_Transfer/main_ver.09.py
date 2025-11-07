# 保存：excel_mapper_v3_3.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
from pathlib import Path

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel転記ツール v3.3")
        self.root.geometry("980x760")

        # 変数
        self.src_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.src_start_row = tk.IntVar(value=2)
        self.src_key_col = tk.StringVar(value="A")

        self.dst_file = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.dst_start_row = tk.IntVar(value=2)
        self.dst_key_col = tk.StringVar(value="A")

        self.overwrite = tk.BooleanVar(value=True)

        self.mappings = []  # list of (src_col_var, dst_col_var, frame)

        self.build_ui()

    def build_ui(self):
        root = self.root
        frame = ttk.Frame(root, padding=10)
        frame.pack(fill="both", expand=True)

        # --- 元データブロック（縦並び） ---
        block_src = ttk.LabelFrame(frame, text="元データ", padding=8)
        block_src.pack(fill="x", padx=4, pady=4)

        row = 0
        ttk.Label(block_src, text="ファイル:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_src, textvariable=self.src_file, width=70).grid(row=row, column=1, sticky="w")
        ttk.Button(block_src, text="選択", command=self.select_src_file).grid(row=row, column=2, padx=4)
        row += 1
        ttk.Label(block_src, text="シート:").grid(row=row, column=0, sticky="w")
        self.src_sheet_cb = ttk.Combobox(block_src, textvariable=self.src_sheet, width=30, state="readonly")
        self.src_sheet_cb.grid(row=row, column=1, sticky="w")
        row += 1
        ttk.Label(block_src, text="データ開始行:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_src, textvariable=self.src_start_row, width=6).grid(row=row, column=1, sticky="w")
        row += 1
        ttk.Label(block_src, text="照合キー列:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_src, textvariable=self.src_key_col, width=6).grid(row=row, column=1, sticky="w")

        # --- 転記設定（横並び：元 -> 先） ---
        map_block = ttk.LabelFrame(frame, text="転記設定（元列 → 転記先列）", padding=8)
        map_block.pack(fill="both", padx=4, pady=6, expand=True)

        self.map_canvas = tk.Canvas(map_block, height=240)
        self.map_canvas.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(map_block, orient="vertical", command=self.map_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        self.map_canvas.configure(yscrollcommand=scrollbar.set)
        self.map_inner = ttk.Frame(self.map_canvas)
        self.map_canvas.create_window((0,0), window=self.map_inner, anchor="nw")
        self.map_inner.bind("<Configure>", lambda e: self.map_canvas.configure(scrollregion=self.map_canvas.bbox("all")))

        # header row for mapping area
        hdr = ttk.Frame(self.map_inner)
        hdr.grid(row=0, column=0, sticky="w", pady=(0,6))
        ttk.Label(hdr, text="元列", width=8).pack(side="left", padx=6)
        ttk.Label(hdr, text="→", width=3).pack(side="left")
        ttk.Label(hdr, text="転記先列", width=12).pack(side="left", padx=6)
        ttk.Label(hdr, text="", width=8).pack(side="left")  # spacer for delete button

        btn_line = ttk.Frame(frame)
        btn_line.pack(fill="x", padx=4)
        ttk.Button(btn_line, text="転記条件追加", command=self.add_mapping).pack(side="left", padx=4)

        # --- 転記先ブロック（縦並び） ---
        block_dst = ttk.LabelFrame(frame, text="転記先", padding=8)
        block_dst.pack(fill="x", padx=4, pady=4)
        row = 0
        ttk.Label(block_dst, text="ファイル:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_dst, textvariable=self.dst_file, width=70).grid(row=row, column=1, sticky="w")
        ttk.Button(block_dst, text="選択", command=self.select_dst_file).grid(row=row, column=2, padx=4)
        row += 1
        ttk.Label(block_dst, text="シート:").grid(row=row, column=0, sticky="w")
        self.dst_sheet_cb = ttk.Combobox(block_dst, textvariable=self.dst_sheet, width=30, state="readonly")
        self.dst_sheet_cb.grid(row=row, column=1, sticky="w")
        row += 1
        ttk.Label(block_dst, text="データ開始行:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_dst, textvariable=self.dst_start_row, width=6).grid(row=row, column=1, sticky="w")
        row += 1
        ttk.Label(block_dst, text="照合キー列:").grid(row=row, column=0, sticky="w")
        ttk.Entry(block_dst, textvariable=self.dst_key_col, width=6).grid(row=row, column=1, sticky="w")

        # --- 下部コントロール ---
        ctrl = ttk.Frame(frame)
        ctrl.pack(fill="x", padx=4, pady=8)
        ttk.Checkbutton(ctrl, text="既存データを上書きする", variable=self.overwrite).pack(side="left")
        ttk.Button(ctrl, text="転記実行", command=self.execute_mapping).pack(side="right", padx=6)

        # 初期1行
        self.add_mapping()

    # mapping 行追加（見た目: 元:[ ] → 先:[ ] [削除]）
    def add_mapping(self):
        idx = len(self.mappings) + 1  # start at row 1 (row 0 is header)
        src_var = tk.StringVar()
        dst_var = tk.StringVar()
        row_frame = ttk.Frame(self.map_inner)
        row_frame.grid(row=idx, column=0, sticky="w", pady=3, padx=2)

        ttk.Entry(row_frame, textvariable=src_var, width=8).pack(side="left", padx=(6,4))
        ttk.Label(row_frame, text="→").pack(side="left", padx=4)
        ttk.Entry(row_frame, textvariable=dst_var, width=12).pack(side="left", padx=(4,10))
        ttk.Button(row_frame, text="削除", command=lambda f=row_frame: self.delete_mapping(f)).pack(side="left", padx=4)

        self.mappings.append((src_var, dst_var, row_frame))

    def delete_mapping(self, frame):
        # find and remove mapping
        for i, (s, d, f) in enumerate(self.mappings):
            if f == frame:
                f.destroy()
                self.mappings.pop(i)
                break
        # re-grid remaining rows to keep order
        for j, (_, _, f) in enumerate(self.mappings, start=1):
            f.grid_configure(row=j)

    # ファイル選択
    def select_src_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.src_file.set(path)
            xl = pd.ExcelFile(path)
            self.src_sheet_cb["values"] = xl.sheet_names
            if xl.sheet_names:
                self.src_sheet.set(xl.sheet_names[0])

    def select_dst_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.dst_file.set(path)
            xl = pd.ExcelFile(path)
            self.dst_sheet_cb["values"] = xl.sheet_names
            if xl.sheet_names:
                self.dst_sheet.set(xl.sheet_names[0])

    # 列文字 -> 0-based index
    def col_to_index(self, col):
        col = col.strip().upper()
        idx = 0
        for c in col:
            if 'A' <= c <= 'Z':
                idx = idx * 26 + (ord(c) - ord('A') + 1)
        return idx - 1

    # 実作業：転記
    def execute_mapping(self):
        try:
            # 基本チェック
            if not (self.src_file.get() and self.dst_file.get() and self.src_sheet.get() and self.dst_sheet.get()):
                messagebox.showerror("エラー", "元・転記先のファイルとシートを選択してください。")
                return
            if not self.mappings:
                messagebox.showerror("エラー", "転記条件を1件以上設定してください。")
                return
            if not (self.src_key_col.get() and self.dst_key_col.get()):
                messagebox.showerror("エラー", "照合キー列（元・転記先）を指定してください。")
                return

            src_path = self.src_file.get()
            dst_path = self.dst_file.get()
            src_sheet = self.src_sheet.get()
            dst_sheet = self.dst_sheet.get()

            # 読み込み（header=Noneで生データとして扱う）
            src_df = pd.read_excel(src_path, sheet_name=src_sheet, header=None, skiprows=self.src_start_row.get()-1)
            dst_df = pd.read_excel(dst_path, sheet_name=dst_sheet, header=None, skiprows=self.dst_start_row.get()-1)

            # Excel上の行番号をインデックスにする（画面左端行番号）
            src_df.index = range(self.src_start_row.get(), self.src_start_row.get() + len(src_df))
            dst_df.index = range(self.dst_start_row.get(), self.dst_start_row.get() + len(dst_df))

            # 必要列数を確認して補完（dst）
            needed_cols = []
            for s_var, d_var, _ in self.mappings:
                try:
                    needed_cols.append(self.col_to_index(s_var.get()))
                except Exception:
                    pass
                try:
                    needed_cols.append(self.col_to_index(d_var.get()))
                except Exception:
                    pass
            max_needed = max(needed_cols) + 1 if needed_cols else 0
            while len(dst_df.columns) < max_needed:
                dst_df[len(dst_df.columns)] = None

            # 照合キーのSeries
            src_key_idx = self.col_to_index(self.src_key_col.get())
            dst_key_idx = self.col_to_index(self.dst_key_col.get())
            key_src = src_df.iloc[:, src_key_idx]
            key_dst = dst_df.iloc[:, dst_key_idx]

            # ログ準備
            desktop = Path.home() / "Desktop"
            if not desktop.exists():
                desktop = Path.home() / "デスクトップ"
            log_dir = desktop / "コピペログ"
            log_dir.mkdir(exist_ok=True)
            now = datetime.now().strftime("%Y.%m.%d.%H.%M.%S")
            log_path = log_dir / f"{now}.xlsx"

            wb_log = Workbook()
            ws_log = wb_log.active
            ws_log["A1"] = "転記元ファイル"
            ws_log["B2"] = Path(src_path).name
            ws_log["B3"] = src_sheet
            ws_log["A4"] = "転記先ファイル"
            ws_log["B5"] = Path(dst_path).name
            ws_log["B6"] = dst_sheet
            ws_log["C7"] = "上書き" if self.overwrite.get() else "既存保持"

            # ログヘッダー組み立て
            headers = ["転記先行", "転記元行", "一致キー"]
            for s_var, d_var, _ in self.mappings:
                headers.append(f"転記先{d_var.get()}(列名)")
                headers.append(f"転記元{s_var.get()}(列名)")
            for i, h in enumerate(headers):
                ws_log.cell(row=8, column=i+1, value=h)
            log_row = 9

            # Prepare destination workbook for cell-level writes to preserve headers & other sheets
            wb_dst = load_workbook(dst_path)
            if dst_sheet not in wb_dst.sheetnames:
                # create sheet
                wb_dst.create_sheet(dst_sheet)
            ws_dst = wb_dst[dst_sheet]

            # 実処理ループ（dstの各行を回す）
            for i_dst, key_val in enumerate(key_dst):
                # 空白ならスキップ
                if pd.isna(key_val) or str(key_val).strip() == "":
                    continue

                # key_val に一致する src_df行を探す
                match_rows = src_df[key_src == key_val]
                if match_rows.empty:
                    continue

                # srcの実行行番号（Excel上の行番号）
                src_row_excel = match_rows.index[0]
                # dstの実行行番号（Excel上の行番号）
                dst_row_excel = dst_df.index[i_dst]

                log_values = [dst_row_excel, src_row_excel, key_val]

                # 各マッピングでセル単位処理（openpyxl を使ってシート上のセルを直接上書き）
                for s_var, d_var, _ in self.mappings:
                    s_idx = self.col_to_index(s_var.get())
                    d_idx = self.col_to_index(d_var.get())
                    # src_val を安全取得（範囲外なら None）
                    src_val = None
                    if s_idx >= 0 and s_idx < len(src_df.columns):
                        src_val = src_df.iat[src_df.index.get_loc(src_row_excel), s_idx]

                    # ワークシート上の既存値取得（列は1-based）
                    cell = ws_dst.cell(row=dst_row_excel, column=d_idx+1)
                    dst_before = cell.value

                    # 上書き or 既存保持 の判断
                    if self.overwrite.get() or dst_before is None:
                        cell.value = src_val
                        dst_after = src_val
                    else:
                        dst_after = dst_before

                    # ログに転記先(最終)と転記元を追加
                    log_values.append(dst_after)
                    log_values.append(src_val)

                    # keep dst_df in sync for any further checks (optional)
                    # ensure dst_df has that column index
                    if d_idx >= len(dst_df.columns):
                        # expand dst_df columns
                        while len(dst_df.columns) <= d_idx:
                            dst_df[len(dst_df.columns)] = None
                    # set value in dst_df row (by position)
                    dst_df.iat[i_dst, d_idx] = dst_after

                # 書き込みログ行追加
                for j, v in enumerate(log_values):
                    ws_log.cell(row=log_row, column=j+1, value=v)
                log_row += 1

            # 保存：ワークブック（セル単位で変更）を保存
            wb_dst.save(dst_path)
            # ログ保存
            wb_log.save(log_path)

            messagebox.showinfo("完了", f"転記完了。\nログ: {log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラーが発生しました:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
