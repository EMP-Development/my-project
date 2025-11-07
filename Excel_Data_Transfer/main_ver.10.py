import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel転記ツール v3.4（確認モード対応）")
        self.root.geometry("1100x700")

        self.src_file = tk.StringVar()
        self.dst_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.src_key_col = tk.StringVar()
        self.dst_key_col = tk.StringVar()
        self.start_row = tk.IntVar(value=1)
        self.exclude_cols = tk.StringVar(value="")
        self.exec_mode = tk.StringVar(value="overwrite")

        self.mappings = []

        self.build_ui()

    def build_ui(self):
        # 上部フレーム：元データと転記先データを左右に配置
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)

        # 元データブロック
        src_frame = ttk.LabelFrame(top_frame, text="元データ")
        src_frame.pack(side="left", fill="both", expand=True, padx=5)

        ttk.Button(src_frame, text="ファイル選択", command=lambda: self.select_file(self.src_file, self.src_sheet)).pack(pady=2)
        ttk.Entry(src_frame, textvariable=self.src_file, width=50).pack(pady=2)
        ttk.Label(src_frame, text="シート選択:").pack()
        self.src_sheet_cb = ttk.Combobox(src_frame, textvariable=self.src_sheet, state="readonly")
        self.src_sheet_cb.pack(pady=2)
        ttk.Label(src_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_key_col, width=10).pack(pady=2)

        # 転記先データブロック
        dst_frame = ttk.LabelFrame(top_frame, text="転記先データ")
        dst_frame.pack(side="right", fill="both", expand=True, padx=5)

        ttk.Button(dst_frame, text="ファイル選択", command=lambda: self.select_file(self.dst_file, self.dst_sheet)).pack(pady=2)
        ttk.Entry(dst_frame, textvariable=self.dst_file, width=50).pack(pady=2)
        ttk.Label(dst_frame, text="シート選択:").pack()
        self.dst_sheet_cb = ttk.Combobox(dst_frame, textvariable=self.dst_sheet, state="readonly")
        self.dst_sheet_cb.pack(pady=2)
        ttk.Label(dst_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_key_col, width=10).pack(pady=2)

        # 転記設定部分
        mapping_frame = ttk.LabelFrame(self.root, text="転記設定（元列 → 転記先列）")
        mapping_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.mapping_container = ttk.Frame(mapping_frame)
        self.mapping_container.pack(fill="x", padx=5, pady=5)

        ttk.Button(mapping_frame, text="＋ 条件を追加", command=self.add_mapping).pack(anchor="e", padx=10, pady=5)
        self.add_mapping()  # 初期1行

        # 下部の設定
        bottom_frame = ttk.LabelFrame(self.root, text="設定・実行")
        bottom_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(bottom_frame, text="開始行:").pack(side="left")
        ttk.Entry(bottom_frame, textvariable=self.start_row, width=5).pack(side="left", padx=(0,10))

        ttk.Label(bottom_frame, text="除外列（例: A,B,C）:").pack(side="left")
        ttk.Entry(bottom_frame, textvariable=self.exclude_cols, width=20).pack(side="left", padx=(0,10))

        ttk.Label(bottom_frame, text="実行モード:").pack(side="left", padx=(10,0))
        ttk.Radiobutton(bottom_frame, text="上書き", value="overwrite", variable=self.exec_mode).pack(side="left")
        ttk.Radiobutton(bottom_frame, text="既存保持", value="keep", variable=self.exec_mode).pack(side="left")
        ttk.Radiobutton(bottom_frame, text="確認のみ", value="preview", variable=self.exec_mode).pack(side="left", padx=(10,0))

        ttk.Button(bottom_frame, text="転記実行", command=self.execute_mapping).pack(side="right", padx=10)

    # ファイル選択処理
    def select_file(self, var, sheet_var):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not file_path:
            return
        var.set(file_path)
        try:
            xl = pd.ExcelFile(file_path)
            if var == self.src_file:
                self.src_sheet_cb["values"] = xl.sheet_names
                if xl.sheet_names:
                    sheet_var.set(xl.sheet_names[0])
            else:
                self.dst_sheet_cb["values"] = xl.sheet_names
                if xl.sheet_names:
                    sheet_var.set(xl.sheet_names[0])
        except Exception as e:
            messagebox.showerror("エラー", f"シートの取得に失敗しました:\n{e}")

    # 転記設定追加
    def add_mapping(self):
        frame = ttk.Frame(self.mapping_container)
        frame.pack(fill="x", pady=2)

        src_col = tk.StringVar()
        dst_col = tk.StringVar()
        self.mappings.append((src_col, dst_col))

        ttk.Label(frame, text="元列:").pack(side="left")
        ttk.Entry(frame, textvariable=src_col, width=10).pack(side="left", padx=5)
        ttk.Label(frame, text="→ 転記先列:").pack(side="left")
        ttk.Entry(frame, textvariable=dst_col, width=10).pack(side="left", padx=5)

        if len(self.mappings) > 1:
            ttk.Button(frame, text="削除", command=lambda f=frame: self.remove_mapping(f)).pack(side="left", padx=5)

    def remove_mapping(self, frame):
        for w in frame.winfo_children():
            if isinstance(w, ttk.Entry):
                var = w.cget("textvariable")
                if var:
                    var = self.root.getvar(var)
        frame.destroy()

    # 転記実行処理
    def execute_mapping(self):
        try:
            src_path = self.src_file.get()
            dst_path = self.dst_file.get()
            src_sheet = self.src_sheet.get()
            dst_sheet = self.dst_sheet.get()
            src_key_col = self.src_key_col.get().strip()
            dst_key_col = self.dst_key_col.get().strip()
            start_row = self.start_row.get()
            exclude_cols = [c.strip().upper() for c in self.exclude_cols.get().split(",") if c.strip()]
            mode = self.exec_mode.get()

            if not all([src_path, dst_path, src_sheet, dst_sheet, src_key_col, dst_key_col]):
                messagebox.showwarning("警告", "ファイル・シート・照合キー列が指定されていません。")
                return

            mappings = [(s.get().strip().upper(), d.get().strip().upper()) for s, d in self.mappings if s.get().strip() and d.get().strip()]
            if not mappings:
                messagebox.showwarning("警告", "転記設定がありません。")
                return

            df_src = pd.read_excel(src_path, sheet_name=src_sheet, header=None)
            df_dst = pd.read_excel(dst_path, sheet_name=dst_sheet, header=None)
            wb_dst = load_workbook(dst_path)
            ws_dst = wb_dst[dst_sheet]

            log_dir = os.path.join(os.path.expanduser("~/Desktop"), "コピペログ")
            os.makedirs(log_dir, exist_ok=True)
            log_name = datetime.now().strftime("%Y.%m.%d.%H.%M.%S.xlsx")
            log_path = os.path.join(log_dir, log_name)
            wb_log = Workbook()
            ws_log = wb_log.active
            ws_log.title = "転記ログ"

            ws_log["A1"], ws_log["B1"] = "転記元ファイル", os.path.basename(src_path)
            ws_log["A2"], ws_log["B2"] = "転記元シート", src_sheet
            ws_log["A3"], ws_log["B3"] = "転記先ファイル", os.path.basename(dst_path)
            ws_log["A4"], ws_log["B4"] = "転記先シート", dst_sheet
            ws_log["A5"], ws_log["B5"] = "モード", mode

            header = ["転記先行", "転記元行", "照合キー値"]
            for s, d in mappings:
                header += [f"転記先{d}", f"転記元{s}"]
            ws_log.append(header)

            # 照合キー列をインデックスに変換
            def col_to_index(col):
                idx = 0
                for c in col:
                    idx = idx * 26 + (ord(c.upper()) - 64)
                return idx - 1

            src_key_idx = col_to_index(src_key_col)
            dst_key_idx = col_to_index(dst_key_col)

            dst_dict = {}
            for i in range(start_row - 1, len(df_dst)):
                val = df_dst.iat[i, dst_key_idx]
                if pd.notna(val):
                    dst_dict[val] = i

            for i in range(start_row - 1, len(df_src)):
                key_val = df_src.iat[i, src_key_idx]
                if pd.isna(key_val):
                    continue
                if key_val in dst_dict:
                    dst_i = dst_dict[key_val]
                    log_row = [dst_i + 1, i + 1, key_val]

                    for s, d in mappings:
                        s_idx = col_to_index(s)
                        d_idx = col_to_index(d)
                        src_val = df_src.iat[i, s_idx] if s_idx < len(df_src.columns) else None
                        dst_cell = ws_dst.cell(row=dst_i + 1, column=d_idx + 1)
                        dst_before = dst_cell.value
                        dst_after = dst_before

                        if mode == "overwrite":
                            dst_after = src_val
                            dst_cell.value = src_val
                        elif mode == "keep":
                            if dst_before is None:
                                dst_after = src_val
                                dst_cell.value = src_val
                        elif mode == "preview":
                            dst_after = src_val if (dst_before is None or dst_before != src_val) else dst_before

                        log_row += [dst_after, src_val]
                    ws_log.append(log_row)

            wb_log.save(log_path)
            if mode != "preview":
                wb_dst.save(dst_path)

            messagebox.showinfo("完了", f"転記が完了しました。\nログを出力しました：\n{log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラーが発生しました:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
