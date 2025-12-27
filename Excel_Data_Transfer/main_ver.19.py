import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel転記ツール v3.14（縦横自動＋複数形式対応）")
        self.root.geometry("1200x750")

        self.src_file = tk.StringVar()
        self.dst_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.src_key_col = tk.StringVar()
        self.dst_key_col = tk.StringVar()
        self.src_start_row = tk.IntVar(value=1)
        self.dst_start_row = tk.IntVar(value=1)
        self.src_exclude_rows = tk.StringVar(value="")
        self.dst_exclude_rows = tk.StringVar(value="")
        self.exec_mode = tk.StringVar(value="上書き")
        self.mappings = []

        self.build_ui()

    def build_ui(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)

        # 元データ設定
        src_frame = ttk.LabelFrame(top_frame, text="元データ設定")
        src_frame.pack(side="left", fill="both", expand=True, padx=5)

        ttk.Button(src_frame, text="ファイル選択", command=lambda: self.select_file(self.src_file, self.src_sheet, True)).pack(pady=2)
        ttk.Entry(src_frame, textvariable=self.src_file, width=50).pack(pady=2)
        ttk.Label(src_frame, text="シート選択:").pack()
        self.src_sheet_cb = ttk.Combobox(src_frame, textvariable=self.src_sheet, state="readonly")
        self.src_sheet_cb.pack(pady=2)
        ttk.Label(src_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_key_col, width=10).pack(pady=2)
        ttk.Label(src_frame, text="開始行:").pack()
        ttk.Entry(src_frame, textvariable=self.src_start_row, width=5).pack(pady=2)
        ttk.Label(src_frame, text="除外行（例: 1,2,3）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_exclude_rows, width=15).pack(pady=2)

        # 転記先設定
        dst_frame = ttk.LabelFrame(top_frame, text="転記先設定")
        dst_frame.pack(side="right", fill="both", expand=True, padx=5)

        ttk.Button(dst_frame, text="ファイル選択", command=lambda: self.select_file(self.dst_file, self.dst_sheet, False)).pack(pady=2)
        ttk.Entry(dst_frame, textvariable=self.dst_file, width=50).pack(pady=2)
        ttk.Label(dst_frame, text="シート選択:").pack()
        self.dst_sheet_cb = ttk.Combobox(dst_frame, textvariable=self.dst_sheet, state="readonly")
        self.dst_sheet_cb.pack(pady=2)
        ttk.Label(dst_frame, text="照合キー列（例: A）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_key_col, width=10).pack(pady=2)
        ttk.Label(dst_frame, text="開始行:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_start_row, width=5).pack(pady=2)
        ttk.Label(dst_frame, text="除外行（例: 1,2,3）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_exclude_rows, width=15).pack(pady=2)

        # 転記設定
        mapping_frame = ttk.LabelFrame(self.root, text="転記設定（入力元 → 出力先）")
        mapping_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.mapping_container = ttk.Frame(mapping_frame)
        self.mapping_container.pack(fill="x", padx=5, pady=5)
        ttk.Button(mapping_frame, text="＋ 条件を追加", command=self.add_mapping).pack(anchor="e", padx=10, pady=5)
        self.add_mapping()

        # 実行設定
        bottom_frame = ttk.LabelFrame(self.root, text="実行設定")
        bottom_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(bottom_frame, text="実行モード:").pack(side="left")
        for label in ("上書き", "既存保持", "確認のみ"):
            ttk.Radiobutton(bottom_frame, text=label, value=label, variable=self.exec_mode).pack(side="left", padx=5)
        ttk.Button(bottom_frame, text="転記実行", command=self.execute_mapping).pack(side="right", padx=10)

    def select_file(self, var, sheet_var, is_src=True):
        file_path = filedialog.askopenfilename(filetypes=[("Excel & CSV", "*.xlsx *.xlsm *.xls *.csv")])
        if not file_path:
            return
        var.set(file_path)
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext in [".xlsx", ".xlsm", ".xls"]:
                xl = pd.ExcelFile(file_path)
                sheets = xl.sheet_names
            else:
                sheets = ["Sheet1"]
            if is_src:
                self.src_sheet_cb["values"] = sheets
                sheet_var.set(sheets[0])
            else:
                self.dst_sheet_cb["values"] = sheets
                sheet_var.set(sheets[0])
        except Exception as e:
            messagebox.showerror("エラー", f"シート取得失敗:\n{e}")

    def add_mapping(self):
        frame = ttk.Frame(self.mapping_container)
        frame.pack(fill="x", pady=2)
        src_col, dst_col = tk.StringVar(), tk.StringVar()
        self.mappings.append((src_col, dst_col))
        ttk.Label(frame, text="入力元列:").pack(side="left")
        ttk.Entry(frame, textvariable=src_col, width=10).pack(side="left", padx=5)
        ttk.Label(frame, text="→ 出力先列:").pack(side="left")
        ttk.Entry(frame, textvariable=dst_col, width=10).pack(side="left", padx=5)
        if len(self.mappings) > 1:
            ttk.Button(frame, text="削除", command=lambda f=frame: self.remove_mapping(f)).pack(side="left", padx=5)

    def remove_mapping(self, frame):
        frame.destroy()

    def col_to_index(self, col):
        idx = 0
        for c in col:
            if 'A' <= c.upper() <= 'Z':
                idx = idx * 26 + (ord(c.upper()) - 64)
        return idx - 1

    def execute_mapping(self):
        try:
            src_path, dst_path = self.src_file.get(), self.dst_file.get()
            src_sheet, dst_sheet = self.src_sheet.get(), self.dst_sheet.get()
            src_key_col, dst_key_col = self.src_key_col.get().strip(), self.dst_key_col.get().strip()
            src_start, dst_start = self.src_start_row.get(), self.dst_start_row.get()
            mode = self.exec_mode.get()
            src_exclude = [int(x.strip()) for x in self.src_exclude_rows.get().split(",") if x.strip()]
            dst_exclude = [int(x.strip()) for x in self.dst_exclude_rows.get().split(",") if x.strip()]

            mappings = [(s.get().strip().upper(), d.get().strip().upper()) for s, d in self.mappings if s.get().strip() and d.get().strip()]
            if not mappings:
                messagebox.showwarning("警告", "転記設定がありません。")
                return

            # ファイル読み込み（Excel / CSV）
            df_src = self.read_file(src_path, src_sheet)
            df_dst = self.read_file(dst_path, dst_sheet, dst=True)
            wb_dst = load_workbook(dst_path) if dst_path.lower().endswith((".xlsx", ".xlsm")) else Workbook()
            ws_dst = wb_dst[dst_sheet] if dst_sheet in wb_dst.sheetnames else wb_dst.active

            # ログ設定
            log_dir = os.path.join(os.path.expanduser("~/Desktop"), "コピペログ")
            os.makedirs(log_dir, exist_ok=True)
            log_path = os.path.join(log_dir, datetime.now().strftime("%Y.%m.%d.%H.%M.%S.xlsx"))
            wb_log = Workbook()
            ws_log = wb_log.active
            ws_log.title = "転記ログ"

            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            ws_log["A1"], ws_log["B2"], ws_log["B3"] = "転記元ファイル", os.path.basename(src_path), src_sheet
            ws_log["A4"], ws_log["B5"], ws_log["B6"] = "転記先ファイル", os.path.basename(dst_path), dst_sheet
            ws_log["C7"] = f"モード: {mode}"

            # ヘッダー作成（入力元 → 出力先）
            header = ["転記先行", "転記元行", "照合キー値"]
            for s, d in mappings:
                header += [f"入力元{s}", f"出力先{d}"]
            for i, h in enumerate(header):
                ws_log.cell(row=8, column=i + 1, value=h)

            src_key_idx = self.col_to_index(src_key_col)
            dst_key_idx = self.col_to_index(dst_key_col)

            # 縦横自動判定
            direction = "vertical" if df_src.shape[0] >= df_src.shape[1] else "horizontal"

            # 転記先キー辞書
            dst_dict = {}
            if direction == "horizontal":
                for i in range(dst_start - 1, len(df_dst)):
                    if (i + 1) in dst_exclude:
                        continue
                    val = df_dst.iat[i, dst_key_idx] if dst_key_idx < len(df_dst.columns) else None
                    if pd.notna(val):
                        dst_dict[val] = i
            else:
                for i in range(dst_start - 1, len(df_dst.columns)):
                    if (i + 1) in dst_exclude:
                        continue
                    val = df_dst.iat[dst_key_idx, i] if dst_key_idx < len(df_dst) else None
                    if pd.notna(val):
                        dst_dict[val] = i

            log_row = 9
            if direction == "horizontal":
                for i in range(src_start - 1, len(df_src)):
                    if (i + 1) in src_exclude:
                        continue
                    key_val = df_src.iat[i, src_key_idx] if src_key_idx < len(df_src.columns) else None
                    if pd.isna(key_val) or key_val not in dst_dict:
                        continue
                    dst_i = dst_dict[key_val]
                    self.process_row_col(df_src, df_dst, mappings, i, dst_i, ws_dst, ws_log, log_row, key_val, yellow_fill, red_fill, mode)
                    log_row += 1
            else:
                for i in range(src_start - 1, len(df_src.columns)):
                    if (i + 1) in src_exclude:
                        continue
                    key_val = df_src.iat[src_key_idx, i] if src_key_idx < len(df_src) else None
                    if pd.isna(key_val) or key_val not in dst_dict:
                        continue
                    dst_i = dst_dict[key_val]
                    self.process_row_col(df_src, df_dst, mappings, i, dst_i, ws_dst, ws_log, log_row, key_val, yellow_fill, red_fill, mode, vertical=True)
                    log_row += 1

            wb_log.save(log_path)
            if mode != "確認のみ":
                wb_dst.save(dst_path)

            messagebox.showinfo("完了", f"転記完了（モード：{mode}）\nログ出力: {log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラー:\n{e}")

    def read_file(self, path, sheet, dst=False):
        ext = os.path.splitext(path)[1].lower()
        if ext in [".xlsx", ".xlsm", ".xls"]:
            return pd.read_excel(path, sheet_name=sheet, header=None)
        elif ext == ".csv":
            return pd.read_csv(path, header=None)
        else:
            raise ValueError("未対応のファイル形式です")

    def process_row_col(self, df_src, df_dst, mappings, src_idx, dst_idx, ws_dst, ws_log, log_row, key_val, yellow_fill, red_fill, mode, vertical=False):
        log_values = [dst_idx + 1 if not vertical else dst_idx + 1, src_idx + 1 if not vertical else src_idx + 1, key_val]
        color_marks = []
        for s, d in mappings:
            s_idx, d_idx = self.col_to_index(s), self.col_to_index(d)
            if vertical:
                src_val = df_src.iat[s_idx, src_idx] if s_idx < len(df_src) else None
                dst_val = df_dst.iat[d_idx, dst_idx] if d_idx < len(df_dst) else None
            else:
                src_val = df_src.iat[src_idx, s_idx] if s_idx < len(df_src.columns) else None
                dst_val = df_dst.iat[dst_idx, d_idx] if d_idx < len(df_dst.columns) else None

            src_empty = pd.isna(src_val) or src_val == ""
            dst_empty = pd.isna(dst_val) or dst_val == ""
            different = src_val != dst_val

            # 書き込み判定
            if mode == "上書き" and different and not src_empty:
                if vertical:
                    ws_dst.cell(row=d_idx + 1, column=dst_idx + 1).value = src_val
                else:
                    ws_dst.cell(row=dst_idx + 1, column=d_idx + 1).value = src_val
            elif mode == "既存保持" and different and dst_empty:
                if vertical:
                    ws_dst.cell(row=d_idx + 1, column=dst_idx + 1).value = src_val
                else:
                    ws_dst.cell(row=dst_idx + 1, column=d_idx + 1).value = src_val

            # 色付け判定
            if src_empty or dst_empty:
                color_type = "yellow"
            elif different:
                color_type = "red"
            else:
                color_type = None

            log_values += [src_val, dst_val]
            color_marks.append(color_type)

        # ログ書き込み
        for col_idx, v in enumerate(log_values, start=1):
            cell = ws_log.cell(row=log_row, column=col_idx, value=v)
            if col_idx > 3:
                idx = (col_idx - 4) // 2
                color = color_marks[idx] if idx < len(color_marks) else None
                if color == "yellow":
                    cell.fill = yellow_fill
                elif color == "red":
                    cell.fill = red_fill


if __name__ == "__main__":
    root = tk.Tk()
    ExcelMapperApp(root)
    root.mainloop()
