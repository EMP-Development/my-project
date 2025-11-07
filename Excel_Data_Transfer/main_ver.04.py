import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
import os
from datetime import datetime


class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel データ抽出・転記ツール（実用版）")
        self.root.geometry("1150x780")

        # UI用変数
        self.src_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.dst_file = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.src_start_row = tk.IntVar(value=1)
        self.dst_start_row = tk.IntVar(value=1)
        self.overwrite = tk.BooleanVar(value=True)

        # 転記設定
        self.mappings = []
        self.skip_columns = []

        self.key_col = "A"  # 固定キー列

        self.build_ui()

    # === UI構築 ===
    def build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        frame = ttk.Frame(self.root, padding=10)
        frame.grid(sticky="nsew")

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

        # --- 元データ ---
        src = ttk.LabelFrame(frame, text="元データ", padding=10)
        src.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        ttk.Label(src, text="ファイル:").grid(row=0, column=0)
        ttk.Entry(src, textvariable=self.src_file).grid(row=0, column=1, sticky="ew")
        ttk.Button(src, text="参照", command=self.select_src_file).grid(row=0, column=2)
        ttk.Label(src, text="シート:").grid(row=1, column=0)
        self.src_combo = ttk.Combobox(src, textvariable=self.src_sheet, state="readonly")
        self.src_combo.grid(row=1, column=1, sticky="ew")
        ttk.Label(src, text="データ開始行:").grid(row=2, column=0)
        ttk.Entry(src, textvariable=self.src_start_row, width=6).grid(row=2, column=1, sticky="w")

        # --- 転記先 ---
        dst = ttk.LabelFrame(frame, text="転記先", padding=10)
        dst.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        ttk.Label(dst, text="ファイル:").grid(row=0, column=0)
        ttk.Entry(dst, textvariable=self.dst_file).grid(row=0, column=1, sticky="ew")
        ttk.Button(dst, text="参照", command=self.select_dst_file).grid(row=0, column=2)
        ttk.Label(dst, text="シート:").grid(row=1, column=0)
        self.dst_combo = ttk.Combobox(dst, textvariable=self.dst_sheet, state="readonly")
        self.dst_combo.grid(row=1, column=1, sticky="ew")
        ttk.Label(dst, text="データ開始行:").grid(row=2, column=0)
        ttk.Entry(dst, textvariable=self.dst_start_row, width=6).grid(row=2, column=1, sticky="w")

        # --- 転記設定 ---
        map_frame = ttk.LabelFrame(frame, text="転記設定（元列 → 先列）", padding=10)
        map_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=10)
        map_frame.columnconfigure(0, weight=1)

        self.mapping_container = ttk.Frame(map_frame)
        self.mapping_container.grid(row=0, column=0, sticky="nsew")
        self.add_mapping_row()

        ttk.Button(map_frame, text="設定を追加", command=self.add_mapping_row).grid(row=1, column=0, sticky="e")
        ttk.Button(map_frame, text="最後の設定を削除", command=self.remove_mapping_row).grid(row=1, column=0, sticky="w", padx=5)

        # --- 実行 ---
        bottom = ttk.Frame(frame)
        bottom.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Radiobutton(bottom, text="上書きする", variable=self.overwrite, value=True).pack(side="left", padx=5)
        ttk.Radiobutton(bottom, text="既存保持", variable=self.overwrite, value=False).pack(side="left", padx=5)
        ttk.Button(bottom, text="転記実行", command=self.execute_mapping).pack(side="right", padx=10)

    # === ファイル選択 ===
    def select_src_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.src_file.set(path)
            self.load_sheets(path, self.src_combo, self.src_sheet)

    def select_dst_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.dst_file.set(path)
            self.load_sheets(path, self.dst_combo, self.dst_sheet)

    def load_sheets(self, path, combo, var):
        try:
            xl = pd.ExcelFile(path)
            combo["values"] = xl.sheet_names
            if xl.sheet_names:
                var.set(xl.sheet_names[0])
        except Exception as e:
            messagebox.showerror("エラー", f"シート名取得失敗:\n{e}")

    # === 転記設定行 ===
    def add_mapping_row(self):
        frame = ttk.Frame(self.mapping_container)
        frame.pack(fill="x", pady=2)
        src = tk.StringVar()
        dst = tk.StringVar()
        ttk.Entry(frame, textvariable=src, width=10).pack(side="left", padx=5)
        ttk.Label(frame, text="→").pack(side="left")
        ttk.Entry(frame, textvariable=dst, width=10).pack(side="left", padx=5)
        ttk.Button(frame, text="削除", command=lambda f=frame: self.delete_mapping(f)).pack(side="right")
        self.mappings.append((src, dst, frame))

    def remove_mapping_row(self):
        if self.mappings:
            _, _, frame = self.mappings.pop()
            frame.destroy()

    def delete_mapping(self, target):
        for i, (_, _, f) in enumerate(self.mappings):
            if f == target:
                f.destroy()
                del self.mappings[i]
                break

    # === 転記実行 ===
    def execute_mapping(self):
        try:
            src_path = self.src_file.get()
            dst_path = self.dst_file.get()
            src_sheet = self.src_sheet.get()
            dst_sheet = self.dst_sheet.get()

            if not all([src_path, dst_path, src_sheet, dst_sheet]):
                messagebox.showerror("エラー", "すべてのファイルとシートを選択してください。")
                return

            # データ読み込み
            src_df = pd.read_excel(src_path, sheet_name=src_sheet, header=None, skiprows=self.src_start_row.get()-1)
            dst_df = pd.read_excel(dst_path, sheet_name=dst_sheet, header=None, skiprows=self.dst_start_row.get()-1)

            key_src = src_df.iloc[:, 0]
            key_dst = dst_df.iloc[:, 0]

            # ログファイル準備
            desktop = os.path.join(os.path.expanduser("~/Desktop"), "コピペログ")
            os.makedirs(desktop, exist_ok=True)
            now = datetime.now().strftime("%Y.%m.%d.%H.%M.%S")
            log_path = os.path.join(desktop, f"{now}.xls")
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "転記元ファイル"
            ws["B2"] = os.path.basename(src_path)
            ws["B3"] = src_sheet
            ws["A4"] = "転記先ファイル"
            ws["B5"] = os.path.basename(dst_path)
            ws["B6"] = dst_sheet
            ws["C7"] = "上書き" if self.overwrite.get() else "既存保持"

            # ヘッダー行生成
            headers = ["転記先行", "転記元行", "一致キー"]
            for i, (s, d, _) in enumerate(self.mappings):
                base_col = 4 + i * 3
                headers += [f"{d.get()}列(転記先列)", "転記先内容", "転記元内容"]
            for i, h in enumerate(headers):
                ws.cell(row=8, column=i+1, value=h)

            log_row = 9
            # === 実際の転記処理 ===
            for i_dst, key_val in enumerate(key_dst):
                match_rows = src_df[key_src == key_val]
                if not match_rows.empty:
                    src_row_idx = match_rows.index[0]
                    log_values = [i_dst + self.dst_start_row.get(), src_row_idx + self.src_start_row.get(), key_val]

                    for s, d, _ in self.mappings:
                        s_idx = self.col_to_index(s.get())
                        d_idx = self.col_to_index(d.get())

                        src_val = src_df.iat[src_row_idx, s_idx]
                        dst_val = dst_df.iat[i_dst, d_idx] if d_idx < len(dst_df.columns) else None

                        if self.overwrite.get() or pd.isna(dst_val):
                            dst_df.iat[i_dst, d_idx] = src_val

                        log_values += [d.get(), dst_df.iat[i_dst, d_idx], src_val]

                    for j, v in enumerate(log_values):
                        ws.cell(row=log_row, column=j+1, value=v)
                    log_row += 1

            # 保存
            dst_df.to_excel(dst_path, sheet_name=dst_sheet, index=False, header=False)
            wb.save(log_path)

            messagebox.showinfo("完了", f"転記完了！\nログを出力しました：\n{log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラーが発生しました:\n{e}")

    def col_to_index(self, col_letter):
        col_letter = col_letter.strip().upper()
        index = 0
        for c in col_letter:
            index = index * 26 + (ord(c) - ord('A') + 1)
        return index - 1


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
