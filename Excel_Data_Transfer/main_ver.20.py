import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel転記ツール 完全版")
        self.root.geometry("1200x750")

        # ファイル・シート・設定
        self.src_file = tk.StringVar()
        self.dst_file = tk.StringVar()
        self.src_sheet = tk.StringVar()
        self.dst_sheet = tk.StringVar()
        self.src_key_col = tk.StringVar()
        self.dst_key_col = tk.StringVar()
        self.src_start_row = tk.IntVar(value=1)
        self.dst_start_row = tk.IntVar(value=1)
        self.src_exclude_rows = tk.StringVar(value="")
        self.dst_exclude_rows = tk.StringVar(value="")
        self.exec_mode = tk.StringVar(value="上書き")
        self.mappings = []

        self.build_ui()

    def build_ui(self):
        # 上段: 左右フレーム
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="both", expand=True, padx=10, pady=5)

        src_frame = ttk.LabelFrame(top_frame, text="元データ設定")
        src_frame.pack(side="left", fill="both", expand=True, padx=5)
        ttk.Button(src_frame, text="ファイル選択", command=lambda: self.select_file(self.src_file, self.src_sheet, True)).pack(pady=2)
        ttk.Entry(src_frame, textvariable=self.src_file, width=50).pack(pady=2)
        ttk.Label(src_frame, text="シート:").pack()
        self.src_sheet_cb = ttk.Combobox(src_frame, textvariable=self.src_sheet, state="readonly")
        self.src_sheet_cb.pack(pady=2)
        ttk.Label(src_frame, text="照合キー列:").pack()
        ttk.Entry(src_frame, textvariable=self.src_key_col, width=5).pack(pady=2)
        ttk.Label(src_frame, text="開始行:").pack()
        ttk.Entry(src_frame, textvariable=self.src_start_row, width=5).pack(pady=2)
        ttk.Label(src_frame, text="除外行（例:1,2,3）:").pack()
        ttk.Entry(src_frame, textvariable=self.src_exclude_rows, width=10).pack(pady=2)

        dst_frame = ttk.LabelFrame(top_frame, text="転記先設定")
        dst_frame.pack(side="left", fill="both", expand=True, padx=5)
        ttk.Button(dst_frame, text="ファイル選択", command=lambda: self.select_file(self.dst_file, self.dst_sheet, False)).pack(pady=2)
        ttk.Entry(dst_frame, textvariable=self.dst_file, width=50).pack(pady=2)
        ttk.Label(dst_frame, text="シート:").pack()
        self.dst_sheet_cb = ttk.Combobox(dst_frame, textvariable=self.dst_sheet, state="readonly")
        self.dst_sheet_cb.pack(pady=2)
        ttk.Label(dst_frame, text="照合キー列:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_key_col, width=5).pack(pady=2)
        ttk.Label(dst_frame, text="開始行:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_start_row, width=5).pack(pady=2)
        ttk.Label(dst_frame, text="除外行（例:1,2,3）:").pack()
        ttk.Entry(dst_frame, textvariable=self.dst_exclude_rows, width=10).pack(pady=2)

        # 中段: 条件設定（縦横スクロール）
        mapping_frame = ttk.LabelFrame(self.root, text="転記設定（入力元 → 出力先）")
        mapping_frame.pack(fill="both", expand=True, padx=10, pady=5)

        canvas = tk.Canvas(mapping_frame)
        canvas.pack(side="left", fill="both", expand=True)

        v_scroll = ttk.Scrollbar(mapping_frame, orient="vertical", command=canvas.yview)
        v_scroll.pack(side="right", fill="y")
        h_scroll = ttk.Scrollbar(mapping_frame, orient="horizontal", command=canvas.xview)
        h_scroll.pack(side="bottom", fill="x")

        canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self.scroll_frame = ttk.Frame(canvas)
        canvas.create_window((0,0), window=self.scroll_frame, anchor="nw")

        self.add_mapping()
        self.add_button_frame = ttk.Frame(mapping_frame)
        self.add_button_frame.place(relx=1.0, rely=1.0, anchor="se")
        ttk.Button(self.add_button_frame, text="＋ 条件を追加", command=self.add_mapping).pack()

        # 下段: 実行モード
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(bottom_frame, text="実行モード:").pack(side="left")
        for label in ("上書き", "既存保持", "確認のみ"):
            ttk.Radiobutton(bottom_frame, text=label, value=label, variable=self.exec_mode).pack(side="left", padx=5)
        ttk.Button(bottom_frame, text="転記実行", command=self.execute_mapping).pack(side="right", padx=10)

    def select_file(self, var, sheet_var, is_src):
        path = filedialog.askopenfilename(filetypes=[("Excel & CSV", "*.xlsx *.xlsm *.xls *.csv")])
        if not path:
            return
        var.set(path)
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in [".xlsx", ".xlsm", ".xls"]:
                xl = pd.ExcelFile(path)
                sheets = xl.sheet_names
            else:
                sheets = ["Sheet1"]
            if is_src:
                self.src_sheet_cb["values"] = sheets
                sheet_var.set(sheets[0])
            else:
                self.dst_sheet_cb["values"] = sheets
                sheet_var.set(sheets[0])
        except Exception as e:
            messagebox.showerror("エラー", f"シート取得失敗:\n{e}")

    def add_mapping(self):
        frame = ttk.Frame(self.scroll_frame)
        frame.pack(fill="x", pady=2)
        src_col, dst_col = tk.StringVar(), tk.StringVar()
        self.mappings.append((src_col, dst_col))

        ttk.Label(frame, text="入力元列:").pack(side="left")
        ttk.Entry(frame, textvariable=src_col, width=10).pack(side="left", padx=5)
        ttk.Label(frame, text="→ 出力先列:").pack(side="left")
        ttk.Entry(frame, textvariable=dst_col, width=10).pack(side="left", padx=5)

        if len(self.mappings) > 1:
            ttk.Button(frame, text="削除", command=lambda f=frame: self.remove_mapping(f)).pack(side="left", padx=5)

    def remove_mapping(self, frame):
        frame.destroy()

    # 実行処理
    def execute_mapping(self):
        try:
            src_path = self.src_file.get()
            dst_path = self.dst_file.get()
            src_sheet = self.src_sheet.get()
            dst_sheet = self.dst_sheet.get()
            mode = self.exec_mode.get()

            if not src_path or not dst_path:
                messagebox.showerror("エラー", "元ファイルと転記先ファイルを選択してください。")
                return
            if not self.mappings:
                messagebox.showerror("エラー", "転記条件を設定してください。")
                return

            # --- 元データ読み込み ---
            src_df = pd.read_excel(src_path, sheet_name=src_sheet, header=None)
            dst_df = pd.read_excel(dst_path, sheet_name=dst_sheet, header=None)

            # キー列と開始行
            src_key_col_idx = int(self.src_key_col.get()) - 1
            dst_key_col_idx = int(self.dst_key_col.get()) - 1
            src_start = self.src_start_row.get() - 1
            dst_start = self.dst_start_row.get() - 1

            # 除外行
            src_exclude = [int(x)-1 for x in self.src_exclude_rows.get().split(",") if x.strip().isdigit()]
            dst_exclude = [int(x)-1 for x in self.dst_exclude_rows.get().split(",") if x.strip().isdigit()]

            # ログ用ワークブック作成
            log_dir = os.path.join(os.path.expanduser("~"), "Desktop", "コピペログ")
            os.makedirs(log_dir, exist_ok=True)
            log_path = os.path.join(log_dir, datetime.now().strftime("%Y.%m.%d.%H.%M.%S.xlsx"))
            wb_log = Workbook()
            ws_log = wb_log.active
            ws_log.title = "転記ログ"
            ws_log.append(["入力元ファイル", os.path.basename(src_path)])
            ws_log.append(["入力元シート", src_sheet])
            ws_log.append(["転記先ファイル", os.path.basename(dst_path)])
            ws_log.append(["転記先シート", dst_sheet])
            ws_log.append(["モード", mode])
            header = []
            for src_col, dst_col in self.mappings:
                header += [f"入力元{src_col.get()}", f"出力先{dst_col.get()}"]
            ws_log.append(header)

            # --- 転記処理 ---
            fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            fill_red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            for dst_idx in range(dst_start, len(dst_df)):
                if dst_idx in dst_exclude:
                    continue
                dst_key_val = dst_df.iat[dst_idx, dst_key_col_idx]
                if pd.isna(dst_key_val):
                    continue

                # 元データ検索
                src_rows = src_df[src_df.iloc[:, src_key_col_idx] == dst_key_val]
                if src_rows.empty:
                    continue
                src_idx = src_rows.index[0]

                row_log = []
                for src_col, dst_col in self.mappings:
                    src_val = src_df.iat[src_idx, int(src_col.get())-1]
                    dst_val = dst_df.iat[dst_idx, int(dst_col.get())-1]

                    row_log += [src_val, dst_val]

                    # 上書きルール判定
                    if pd.isna(src_val):
                        if mode == "上書き":
                            pass  # 書き換えなし
                        elif mode == "既存保持":
                            pass
                        elif mode == "確認のみ":
                            pass
                    else:
                        if src_val != dst_val:
                            if mode == "上書き":
                                dst_df.iat[dst_idx, int(dst_col.get())-1] = src_val
                            elif mode == "既存保持":
                                pass
                            # 確認のみは書き換えなし
                ws_log.append(row_log)

            # 保存
            dst_df.to_excel(dst_path, sheet_name=dst_sheet, index=False, header=False)
            wb_log.save(log_path)
            messagebox.showinfo("完了", f"転記完了！\nログを出力しました：\n{log_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"転記中にエラーが発生しました:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
